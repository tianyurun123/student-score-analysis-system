from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q, Avg, Max, Min, Count
from django.utils import timezone
from django.http import HttpResponse
import json
import logging

from apps.scores.models import Score, ScoreImportLog, ScoreAdjustment, Gradebook
from apps.scores.serializers import (
    ScoreSerializer, ScoreCreateSerializer, ScoreUpdateSerializer,
    ScoreImportLogSerializer, ScoreAdjustmentSerializer,
    ExcelUploadSerializer, ScoreStatisticsSerializer,
    GradebookSerializer, GradebookCreateSerializer
)
from apps.scores.excel_handlers import ExcelScoreHandler
from apps.courses.models import CourseClass, GradingPolicy
from apps.users.models import User
from utils.calculator import ScoreCalculator
from apps.scores.permissions import IsTeacherOrAdmin

logger = logging.getLogger(__name__)


class ScoreViewSet(viewsets.ModelViewSet):
    """成绩视图集"""
    permission_classes = [IsAuthenticated, IsTeacherOrAdmin]
    queryset = Score.objects.all()

    def get_serializer_class(self):
        if self.action == 'create':
            return ScoreCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return ScoreUpdateSerializer
        return ScoreSerializer

    def get_queryset(self):
        """根据用户权限过滤查询集"""
        queryset = super().get_queryset()
        user = self.request.user

        # 如果是学生，只能查看自己的成绩
        if user.user_type == 'student':
            queryset = queryset.filter(student=user)
        # 如果是教师，只能查看自己授课的课程成绩
        elif user.user_type == 'teacher':
            queryset = queryset.filter(
                Q(course_class__main_teacher=user) |
                Q(course_class__assistant_teachers=user) |
                Q(course_class__course__teachers=user)
            ).distinct()

        # 过滤参数
        course_class_id = self.request.query_params.get('course_class_id')
        if course_class_id:
            queryset = queryset.filter(course_class_id=course_class_id)

        student_id = self.request.query_params.get('student_id')
        if student_id:
            queryset = queryset.filter(student__employee_id=student_id)

        is_published = self.request.query_params.get('is_published')
        if is_published is not None:
            queryset = queryset.filter(is_published=is_published.lower() == 'true')

        return queryset.select_related('student', 'course_class', 'course_class__course', 'grading_policy')

    def perform_create(self, serializer):
        """创建成绩时设置创建人，并确保学生已加入班级"""
        score = serializer.save(
            created_by=self.request.user,
            updated_by=self.request.user
        )
        
        # 确保学生已加入该班级（创建Enrollment记录）
        from apps.courses.models import Enrollment
        enrollment, created = Enrollment.objects.get_or_create(
            student=score.student,
            course_class=score.course_class,
            defaults={'is_active': True}
        )
        
        if created:
            logger.info(f"为新增成绩记录自动创建了选课记录: student={score.student.id}, class={score.course_class.id}")
        
        # 计算成绩
        try:
            score.calculate_grade()
            score.save()
        except Exception as e:
            logger.warning(f"计算新成绩记录失败: {str(e)}")

    def perform_update(self, serializer):
        """更新成绩时设置更新人"""
        serializer.save(updated_by=self.request.user)
    
    def list(self, request, *args, **kwargs):
        """列表查询，序列化器会实时计算usual_entry和final_entry
        同时检查是否有新增的学生（有Enrollment但没有Score），如果有则创建Score记录
        """
        course_class_id = request.query_params.get('course_class_id')
        
        # 如果指定了course_class_id，检查是否有新增的学生需要创建成绩记录
        if course_class_id:
            try:
                course_class = CourseClass.objects.get(id=course_class_id)
                # 获取该班级的所有学生（Enrollment）
                from apps.courses.models import Enrollment
                enrollments = Enrollment.objects.filter(
                    course_class=course_class,
                    is_active=True
                ).select_related('student')
                
                # 获取该班级已有的成绩记录
                existing_scores = Score.objects.filter(
                    course_class=course_class
                ).values_list('student_id', flat=True)
                
                # 找出没有成绩记录的学生
                students_without_scores = [
                    enrollment.student for enrollment in enrollments
                    if enrollment.student.id not in existing_scores
                ]
                
                # 为新增的学生创建成绩记录
                if students_without_scores:
                    # 获取评分政策
                    grading_policy = course_class.course.grading_policy if hasattr(course_class.course, 'grading_policy') else None
                    
                    new_scores = []
                    for student in students_without_scores:
                        score = Score.objects.create(
                            student=student,
                            course_class=course_class,
                            grading_policy=grading_policy,
                            created_by=request.user,
                            updated_by=request.user
                        )
                        new_scores.append(score)
                    
                    if new_scores:
                        logger.info(f"为班级 {course_class_id} 的 {len(new_scores)} 名新增学生创建了成绩记录")
                        
                        # 批量计算新创建的成绩（如果需要）
                        for score in new_scores:
                            try:
                                score.calculate_grade()
                                score.save()
                            except Exception as e:
                                logger.warning(f"计算新成绩记录 {score.id} 失败: {str(e)}")
            
            except CourseClass.DoesNotExist:
                pass
            except Exception as e:
                logger.error(f"检查新增学生时出错: {str(e)}")
        
        # 序列化器已经可以实时计算这些字段，不需要在这里更新
        # 为了性能，可以异步批量更新数据库中的值（可选）
        return super().list(request, *args, **kwargs)
    
    @action(detail=False, methods=['post'], url_path='batch-recalculate-entries')
    def batch_recalculate_entries(self, request):
        """批量重新计算平时录入和期末录入（后台任务）"""
        course_class_id = request.data.get('course_class_id')
        if not course_class_id:
            return Response({'error': '请提供course_class_id'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            queryset = self.get_queryset().filter(course_class_id=course_class_id)
            scores_needing_update = queryset.filter(
                Q(usual_entry__isnull=True) | Q(final_entry__isnull=True)
            )
            
            updated_count = 0
            batch_size = 100
            scores_to_update = []
            
            for score in scores_needing_update:
                try:
                    score.calculate_grade()
                    scores_to_update.append(score)
                    
                    # 每批更新一次
                    if len(scores_to_update) >= batch_size:
                        Score.objects.bulk_update(
                            scores_to_update,
                            ['usual_entry', 'final_entry', 'final_grade', 'grade_point', 'grade_level'],
                            batch_size=batch_size
                        )
                        updated_count += len(scores_to_update)
                        scores_to_update = []
                except Exception as e:
                    logger.error(f"重新计算成绩失败 {score.id}: {str(e)}")
            
            # 更新剩余的记录
            if scores_to_update:
                Score.objects.bulk_update(
                    scores_to_update,
                    ['usual_entry', 'final_entry', 'final_grade', 'grade_point', 'grade_level'],
                    batch_size=batch_size
                )
                updated_count += len(scores_to_update)
            
            return Response({
                'message': f'成功重新计算 {updated_count} 条成绩记录的平时录入和期末录入',
                'updated_count': updated_count
            })
        except Exception as e:
            logger.error(f"批量重新计算失败: {str(e)}")
            return Response({'error': f'批量重新计算失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """导入Excel成绩"""
        serializer = ExcelUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        file = serializer.validated_data['file']
        course_class_id = serializer.validated_data['course_class_id']
        column_mapping = serializer.validated_data.get('column_mapping', {})

        try:
            # 解析Excel
            handler = ExcelScoreHandler()
            result = handler.parse_excel(file, course_class_id)

            if not result['success']:
                return Response({
                    'success': False,
                    'message': 'Excel解析失败',
                    'errors': result['errors'],
                    'suggested_fields': result['suggested_fields'],
                }, status=status.HTTP_400_BAD_REQUEST)

            # 创建导入日志
            import_log = ScoreImportLog.objects.create(
                file_name=file.name,
                file_path='',
                course_class_id=course_class_id,
                imported_by=request.user,
                total_rows=len(result['data']),
                column_mapping=result['column_mapping'],
                status='processing'
            )

            # 导入数据
            import_result = handler.import_scores(
                result['data'],
                course_class_id,
                request.user.id,
                result['column_mapping']
            )

            # 更新导入日志
            import_log.success_rows = import_result['success']
            import_log.failed_rows = import_result['failed']
            import_log.error_log = '\n'.join(import_result['errors'])
            import_log.status = 'completed' if import_result['failed'] == 0 else 'partial'
            import_log.completed_at = timezone.now()
            import_log.save()

            return Response({
                'success': True,
                'message': f'导入完成：成功 {import_result["success"]} 条，失败 {import_result["failed"]} 条',
                'data': {
                    'success_count': import_result['success'],
                    'failed_count': import_result['failed'],
                    'errors': import_result['errors'][:10],  # 只返回前10个错误
                },
                'import_log_id': import_log.id,
            })

        except Exception as e:
            logger.error(f"导入Excel失败: {str(e)}")
            return Response({
                'success': False,
                'message': f'导入失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='preview-excel')
    def preview_excel(self, request):
        """预览Excel文件（不导入）"""
        if 'file' not in request.FILES:
            return Response({
                'success': False,
                'error': '请上传文件',
                'errors': ['请上传Excel文件']
            }, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        
        # 检查文件扩展名
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({
                'success': False,
                'error': '文件格式不正确',
                'errors': ['只支持.xlsx和.xls格式的Excel文件']
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            handler = ExcelScoreHandler()
            result = handler.parse_excel(file)
            
            # 如果解析失败，返回更详细的错误信息
            if not result['success']:
                error_msg = '文件解析失败，请检查文件格式'
                if result.get('errors'):
                    error_msg += f"。错误：{'; '.join(result['errors'][:3])}"
                
                return Response({
                    'success': False,
                    'error': error_msg,
                    'data': [],
                    'total_count': 0,
                    'columns': result.get('columns', []),
                    'column_mapping': result.get('column_mapping', {}),
                    'suggested_fields': result.get('suggested_fields', []),
                    'errors': result.get('errors', []),
                })
            
            return Response({
                'success': True,
                'data': result['data'][:20],  # 只返回前20条预览
                'total_count': len(result['data']),
                'columns': result['columns'],
                'column_mapping': result['column_mapping'],
                'suggested_fields': result['suggested_fields'],
                'errors': result['errors'][:10],  # 只返回前10个错误
            })
        except Exception as e:
            logger.error(f"预览Excel文件失败: {str(e)}", exc_info=True)
            return Response({
                'success': False,
                'error': f'预览失败: {str(e)}',
                'errors': [f'预览失败: {str(e)}']
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        """发布成绩"""
        score = self.get_object()
        score.is_published = True
        score.published_at = timezone.now()
        score.save()

        return Response({
            'message': '成绩已发布',
            'published_at': score.published_at
        })

    @action(detail=True, methods=['post'])
    def unpublish(self, request, pk=None):
        """取消发布成绩"""
        score = self.get_object()
        score.is_published = False
        score.published_at = None
        score.save()

        return Response({'message': '成绩已取消发布'})

    @action(detail=False, methods=['get'], url_path='statistics')
    def statistics(self, request):
        """获取成绩统计信息"""
        course_class_id = request.query_params.get('course_class_id')
        if not course_class_id:
            return Response({'error': '请提供course_class_id'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = self.get_queryset().filter(course_class_id=course_class_id)
        scores = [s.final_grade for s in queryset if s.final_grade is not None]

        stats = ScoreCalculator.calculate_statistics(scores)
        serializer = ScoreStatisticsSerializer(stats)

        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """导出成绩为Excel（标准格式）"""
        import pandas as pd
        from io import BytesIO

        queryset = self.get_queryset()
        course_class_id = request.query_params.get('course_class_id')
        if course_class_id:
            queryset = queryset.filter(course_class_id=course_class_id)

        # 构建数据
        data = []
        for score in queryset:
            row = {
                '学号': score.student.employee_id or score.student.username,
                '姓名': score.student.first_name,
                '考勤': score.attendance_score,
                '作业': score.homework_score,
                '实验': score.experiment_score,
                '复习笔记': score.review_note_score,
                '期末': score.final_score,
                '平时总分': score.usual_total,
                '期末总分': score.final_total,
                '最终成绩': score.final_grade,
                '绩点': score.grade_point,
                '等级': score.grade_level,
            }
            # 添加额外成绩字段
            if score.extra_scores:
                row.update(score.extra_scores)
            data.append(row)

        # 创建DataFrame
        df = pd.DataFrame(data)

        # 导出为Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='成绩表')

        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="scores.xlsx"'
        return response

    @action(detail=False, methods=['get'], url_path='classes-with-scores')
    def classes_with_scores(self, request):
        """获取有成绩的班级列表"""
        # 获取有成绩的班级
        queryset = CourseClass.objects.filter(
            scores__isnull=False
        ).distinct().select_related('course', 'main_teacher')
        
        # 根据用户权限过滤
        user = request.user
        if user.user_type == 'teacher':
            queryset = queryset.filter(
                Q(main_teacher=user) |
                Q(assistant_teachers=user) |
                Q(course__teachers=user)
            ).distinct()
        
        # 统计每个班级的成绩数量
        classes_data = []
        for course_class in queryset:
            score_count = Score.objects.filter(course_class=course_class).count()
            classes_data.append({
                'id': course_class.id,
                'course_code': course_class.course.course_code,
                'course_name': course_class.course.course_name,
                'class_name': course_class.class_name,
                'main_teacher_name': course_class.main_teacher.first_name if course_class.main_teacher else '',
                'score_count': score_count,
                'students_count': course_class.enrollment_set.filter(is_active=True).count()
            })
        
        return Response({
            'results': classes_data,
            'count': len(classes_data)
        })

    @action(detail=False, methods=['get'], url_path='export-achievement')
    def export_achievement(self, request):
        """导出课程目标达成度计算表（目标格式）"""
        from io import BytesIO
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from apps.courses.models import CourseClass, CourseObjectiveAchievement
        from utils.objective_calculator import ObjectiveCalculator

        course_class_id = request.query_params.get('course_class_id')
        if not course_class_id:
            return Response({'error': '请提供course_class_id'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            course_class = CourseClass.objects.select_related('course').get(id=course_class_id)
        except CourseClass.DoesNotExist:
            return Response({'error': '课程班级不存在'}, status=status.HTTP_404_NOT_FOUND)

        # 获取成绩数据
        scores = self.get_queryset().filter(course_class_id=course_class_id).select_related(
            'student', 'student__student_profile'
        ).prefetch_related('objective_achievements')

        if not scores.exists():
            return Response({'error': '该班级没有成绩数据'}, status=status.HTTP_400_BAD_REQUEST)

        # 创建Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = course_class.class_name

        # 设置样式
        header_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 第一行：基本信息（先赋值，再合并）
        # 院系专业
        cell_a1 = ws.cell(row=1, column=1, value='院系专业')
        cell_a1.font = header_font
        cell_a1.alignment = center_align
        ws.merge_cells('A1:B1')
        
        # 开课学院
        cell_c1 = ws.cell(row=1, column=3, value=course_class.course.department or '计算机学院')
        cell_c1.alignment = center_align
        ws.merge_cells('C1:E1')
        
        # 课程名称
        cell_f1 = ws.cell(row=1, column=6, value='课程名称：' + course_class.course.course_name)
        cell_f1.font = header_font
        cell_f1.alignment = center_align
        ws.merge_cells('F1:J1')
        
        # 专业班级
        cell_k1 = ws.cell(row=1, column=11, value='专业班级：' + course_class.class_name)
        cell_k1.alignment = center_align
        ws.merge_cells('K1:O1')

        # 第二行：表头（合并单元格的标题）
        # 学号、姓名
        ws.cell(row=2, column=1, value='学号').font = header_font
        ws.cell(row=2, column=1).alignment = center_align
        ws.cell(row=2, column=1).border = border
        ws.cell(row=2, column=2, value='姓名').font = header_font
        ws.cell(row=2, column=2).alignment = center_align
        ws.cell(row=2, column=2).border = border
        
        # 平时成绩（合并C2:E2，包含点名、电子笔记、作业成绩）
        ws.cell(row=2, column=3, value='平时成绩').font = header_font
        ws.cell(row=2, column=3).alignment = center_align
        ws.cell(row=2, column=3).border = border
        
        # 平时成绩（整数，F2）
        ws.cell(row=2, column=6, value='平时成绩').font = header_font
        ws.cell(row=2, column=6).alignment = center_align
        ws.cell(row=2, column=6).border = border
        
        # 实验（G2）
        ws.cell(row=2, column=7, value='实验').font = header_font
        ws.cell(row=2, column=7).alignment = center_align
        ws.cell(row=2, column=7).border = border
        
        # 期末作品（合并H2:J2，包含作品、报告、期末平均）
        ws.cell(row=2, column=8, value='期末作品').font = header_font
        ws.cell(row=2, column=8).alignment = center_align
        ws.cell(row=2, column=8).border = border
        
        # 课程目标1（合并K2:O2，包含平时、实验、期末、达成情况、达成度）
        ws.cell(row=2, column=11, value='课程目标1').font = header_font
        ws.cell(row=2, column=11).alignment = center_align
        ws.cell(row=2, column=11).border = border
        
        # 课程目标2（合并P2:T2，包含平时、实验、期末、达成情况、达成度）
        ws.cell(row=2, column=16, value='课程目标2').font = header_font
        ws.cell(row=2, column=16).alignment = center_align
        ws.cell(row=2, column=16).border = border
        
        # 课程目标3（合并U2:Y2，包含平时、实验、期末、达成情况、达成度）
        ws.cell(row=2, column=21, value='课程目标3').font = header_font
        ws.cell(row=2, column=21).alignment = center_align
        ws.cell(row=2, column=21).border = border
        
        # 总分值（Z2）
        ws.cell(row=2, column=26, value='总分值').font = header_font
        ws.cell(row=2, column=26).alignment = center_align
        ws.cell(row=2, column=26).border = border
        
        # 平时录入（AA2）
        ws.cell(row=2, column=27, value='平时录入').font = header_font
        ws.cell(row=2, column=27).alignment = center_align
        ws.cell(row=2, column=27).border = border
        
        # 期末录入（AB2）
        ws.cell(row=2, column=28, value='期末录入').font = header_font
        ws.cell(row=2, column=28).alignment = center_align
        ws.cell(row=2, column=28).border = border
        
        # 最终成绩（AC2）
        ws.cell(row=2, column=29, value='最终成绩').font = header_font
        ws.cell(row=2, column=29).alignment = center_align
        ws.cell(row=2, column=29).border = border

        # 第三行：详细表头（必须与数据行的列数完全一致）
        headers_row3 = ['', '', '点名', '电子笔记', '作业成绩', '', '', '作品', '报告', '期末平均',
                       '平时', '实验', '期末', '达成情况', '达成度',
                       '平时', '实验', '期末', '达成情况', '达成度',
                       '平时', '实验', '期末', '达成情况', '达成度', '', '', '', '']
        for col_idx, header in enumerate(headers_row3, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # 合并单元格（先赋值，再合并）
        # 确保所有单元格都已赋值后再合并
        ws.merge_cells('A2:A3')  # 学号
        ws.merge_cells('B2:B3')  # 姓名
        ws.merge_cells('C2:E2')  # 平时成绩（点名、电子笔记、作业成绩）
        ws.merge_cells('F2:F3')  # 平时成绩（整数）
        ws.merge_cells('G2:G3')  # 实验
        ws.merge_cells('H2:J2')  # 期末作品（作品、报告、期末平均）
        ws.merge_cells('K2:O2')  # 课程目标1（平时、实验、期末、达成情况、达成度）
        ws.merge_cells('P2:T2')  # 课程目标2（平时、实验、期末、达成情况、达成度）
        ws.merge_cells('U2:Y2')  # 课程目标3（平时、实验、期末、达成情况、达成度）
        ws.merge_cells('Z2:Z3')  # 总分值
        ws.merge_cells('AA2:AA3')  # 平时录入
        ws.merge_cells('AB2:AB3')  # 期末录入
        ws.merge_cells('AC2:AC3')  # 最终成绩
        

        # 填充数据
        row_num = 4
        for score in scores.order_by('student__employee_id', 'student__username'):
            # 确保重新计算成绩（包括usual_entry和final_entry）
            score.calculate_grade()
            
            # 确保计算了课程目标达成度
            achievements = CourseObjectiveAchievement.objects.filter(score=score).order_by('objective_number')
            if not achievements.exists():
                ObjectiveCalculator.save_objective_achievements(score)
                achievements = CourseObjectiveAchievement.objects.filter(score=score).order_by('objective_number')
            
            obj1 = achievements.filter(objective_number=1).first()
            obj2 = achievements.filter(objective_number=2).first()
            obj3 = achievements.filter(objective_number=3).first()

            # 获取成绩数据
            attendance = score.attendance_score or 0
            e_notes = score.extra_scores.get('电子笔记', 0) or 0
            homework = score.homework_score or 0
            experiment = score.experiment_score or 0
            work = score.extra_scores.get('作品', 0) or 0
            report = score.extra_scores.get('报告', 0) or 0

            # 计算平时成绩和期末平均（取整数）
            usual_total = round(score.usual_total) if score.usual_total is not None else 0
            final_total = int(score.final_total) if score.final_total is not None else 0  # 期末平均直接取整，不四舍五入
            
            # 计算总分值 = 课程目标1达成情况 + 课程目标2达成情况 + 课程目标3达成情况
            total_score = (obj1.achievement_score if obj1 else 0) + (obj2.achievement_score if obj2 else 0) + (obj3.achievement_score if obj3 else 0)
            
            # 使用计算后的值
            usual_entry = score.usual_entry if score.usual_entry is not None else 0
            final_entry = score.final_entry if score.final_entry is not None else 0

            # 填充数据（严格按照表头列顺序）
            # 填充数据（严格按照表头第三行的列顺序，共29列）
            # 表头第三行：'', '', '点名', '电子笔记', '作业成绩', '', '', '作品', '报告', '期末平均',
            #            '平时', '实验', '期末', '达成情况', '达成度', (课程目标1)
            #            '平时', '实验', '期末', '达成情况', '达成度', (课程目标2)
            #            '平时', '实验', '期末', '达成情况', '达成度', '', '', '', '' (课程目标3)
            data_row = [
                '',  # A1: 空（学号在第二行）
                '',  # B2: 空（姓名在第二行）
                attendance,  # C3: 点名
                e_notes,  # D4: 电子笔记
                homework,  # E5: 作业成绩
                '',  # F6: 空（平时成绩在第二行）
                '',  # G7: 空（实验在第二行）
                work,  # H8: 作品
                report,  # I9: 报告
                final_total,  # J10: 期末平均（整数，在报告右侧）
                # 课程目标1（K11-O15）
                round(obj1.usual_score, 2) if obj1 else 0.00,  # K11: 平时
                round(obj1.experiment_score, 2) if obj1 else 0.00,  # L12: 实验
                round(obj1.final_score, 2) if obj1 else 0.00,  # M13: 期末
                round(obj1.achievement_score, 2) if obj1 else 0.00,  # N14: 达成情况
                round(obj1.achievement_degree, 2) if obj1 else 0.00,  # O15: 达成度
                # 课程目标2（P16-T20）
                round(obj2.usual_score, 2) if obj2 else 0.00,  # P16: 平时
                round(obj2.experiment_score, 2) if obj2 else 0.00,  # Q17: 实验
                round(obj2.final_score, 2) if obj2 else 0.00,  # R18: 期末
                round(obj2.achievement_score, 2) if obj2 else 0.00,  # S19: 达成情况
                round(obj2.achievement_degree, 2) if obj2 else 0.00,  # T20: 达成度
                # 课程目标3（U21-Y25）
                round(obj3.usual_score, 2) if obj3 else 0.00,  # U21: 平时
                round(obj3.experiment_score, 2) if obj3 else 0.00,  # V22: 实验
                round(obj3.final_score, 2) if obj3 else 0.00,  # W23: 期末
                round(obj3.achievement_score, 2) if obj3 else 0.00,  # X24: 达成情况
                round(obj3.achievement_degree, 2) if obj3 else 0.00,  # Y25: 达成度
                '',  # Z26: 空（总分值在第二行）
                '',  # AA27: 空（平时录入在第二行）
                '',  # AB28: 空（期末录入在第二行）
                '',  # AC29: 空（最终成绩在第二行）
            ]

            # 先填充第三行对应的数据（data_row）
            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.alignment = center_align
                cell.border = border
            
            # 再单独填充第二行合并单元格对应的值（这些列在第三行是空的）
            ws.cell(row=row_num, column=1, value=score.student.employee_id or score.student.username).alignment = center_align
            ws.cell(row=row_num, column=1).border = border
            ws.cell(row=row_num, column=2, value=score.student.first_name or score.student.username).alignment = center_align
            ws.cell(row=row_num, column=2).border = border
            ws.cell(row=row_num, column=6, value=usual_total).alignment = center_align
            ws.cell(row=row_num, column=6).border = border
            ws.cell(row=row_num, column=7, value=experiment).alignment = center_align
            ws.cell(row=row_num, column=7).border = border
            ws.cell(row=row_num, column=26, value=round(total_score, 2)).alignment = center_align
            ws.cell(row=row_num, column=26).border = border
            ws.cell(row=row_num, column=27, value=int(usual_entry) if usual_entry is not None else 0).alignment = center_align
            ws.cell(row=row_num, column=27).border = border
            ws.cell(row=row_num, column=28, value=int(final_entry) if final_entry is not None else 0).alignment = center_align
            ws.cell(row=row_num, column=28).border = border
            ws.cell(row=row_num, column=29, value=int(score.final_grade) if score.final_grade is not None else 0).alignment = center_align
            ws.cell(row=row_num, column=29).border = border

            row_num += 1

        # 在表格最下面添加达成度计算表
        stats_start_row = row_num + 2  # 空一行
        
        # 标题行："达成度计算"居中放在表头正上方
        title_cell = ws.cell(row=stats_start_row, column=2, value='达成度计算')
        title_cell.font = header_font
        title_cell.alignment = center_align
        ws.merge_cells(f'B{stats_start_row}:D{stats_start_row}')  # 合并B、C、D列，居中显示
        
        # 表头行（在标题下方）
        header_row = stats_start_row + 1
        ws.cell(row=header_row, column=2, value='课程目标').font = header_font
        ws.cell(row=header_row, column=2).alignment = center_align
        ws.cell(row=header_row, column=2).border = border
        ws.cell(row=header_row, column=3, value='达成分值').font = header_font
        ws.cell(row=header_row, column=3).alignment = center_align
        ws.cell(row=header_row, column=3).border = border
        ws.cell(row=header_row, column=4, value='达成度').font = header_font
        ws.cell(row=header_row, column=4).alignment = center_align
        ws.cell(row=header_row, column=4).border = border
        
        # 计算班级平均达成度（与前端计算方式一致）
        # 计算每个课程目标的平均达成度和达成分值
        obj1_scores = []
        obj1_degrees = []
        obj2_scores = []
        obj2_degrees = []
        obj3_scores = []
        obj3_degrees = []
        
        for score in scores:
            achievements = CourseObjectiveAchievement.objects.filter(score=score).order_by('objective_number')
            obj1 = achievements.filter(objective_number=1).first()
            obj2 = achievements.filter(objective_number=2).first()
            obj3 = achievements.filter(objective_number=3).first()
            
            if obj1:
                obj1_scores.append(obj1.achievement_score)
                obj1_degrees.append(obj1.achievement_degree)
            if obj2:
                obj2_scores.append(obj2.achievement_score)
                obj2_degrees.append(obj2.achievement_degree)
            if obj3:
                obj3_scores.append(obj3.achievement_score)
                obj3_degrees.append(obj3.achievement_degree)
        
        # 计算平均值（保留两位小数）
        obj1_avg_score = round(sum(obj1_scores) / len(obj1_scores), 2) if obj1_scores else 0.00
        obj1_avg_degree = round(sum(obj1_degrees) / len(obj1_degrees), 2) if obj1_degrees else 0.00  # 不乘以100，保持小数格式
        obj2_avg_score = round(sum(obj2_scores) / len(obj2_scores), 2) if obj2_scores else 0.00
        obj2_avg_degree = round(sum(obj2_degrees) / len(obj2_degrees), 2) if obj2_degrees else 0.00  # 不乘以100，保持小数格式
        obj3_avg_score = round(sum(obj3_scores) / len(obj3_scores), 2) if obj3_scores else 0.00
        obj3_avg_degree = round(sum(obj3_degrees) / len(obj3_degrees), 2) if obj3_degrees else 0.00  # 不乘以100，保持小数格式
        
        # 填充达成度计算数据（与前端界面一致：达成分值保留两位小数，达成度保留两位小数，不显示百分号）
        stats_data = [
            ('课程目标1', obj1_avg_score, obj1_avg_degree),
            ('课程目标2', obj2_avg_score, obj2_avg_degree),
            ('课程目标3', obj3_avg_score, obj3_avg_degree),
        ]
        
        for idx, (objective, score, degree) in enumerate(stats_data, 1):
            row = header_row + idx
            ws.cell(row=row, column=1, value='').border = border
            ws.cell(row=row, column=2, value=objective).border = border
            ws.cell(row=row, column=2).alignment = center_align
            ws.cell(row=row, column=3, value=score).border = border
            ws.cell(row=row, column=3).alignment = center_align
            ws.cell(row=row, column=4, value=degree).border = border  # 不添加%符号，直接显示小数
            ws.cell(row=row, column=4).alignment = center_align

        # 设置列宽
        column_widths = {
            'A': 15, 'B': 10, 'C': 8, 'D': 10, 'E': 10, 'F': 10, 'G': 8,
            'H': 8, 'I': 8, 'J': 10, 'K': 5,
            'L': 8, 'M': 8, 'N': 8, 'O': 10, 'P': 10,
            'Q': 8, 'R': 8, 'S': 8, 'T': 10, 'U': 10,
            'V': 8, 'W': 8, 'X': 8, 'Y': 10, 'Z': 10,
            'AA': 10, 'AB': 10, 'AC': 10, 'AD': 10
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{course_class.course.course_name}_{course_class.class_name}_达成度计算.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ScoreImportLogViewSet(viewsets.ReadOnlyModelViewSet):
    """成绩导入日志视图集"""
    permission_classes = [IsAuthenticated, IsTeacherOrAdmin]
    queryset = ScoreImportLog.objects.all()
    serializer_class = ScoreImportLogSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user

        if user.user_type == 'teacher':
            queryset = queryset.filter(
                Q(course_class__main_teacher=user) |
                Q(course_class__assistant_teachers=user) |
                Q(imported_by=user)
            ).distinct()

        return queryset.select_related('course_class', 'course_class__course', 'imported_by')


class ScoreAdjustmentViewSet(viewsets.ModelViewSet):
    """成绩调整记录视图集"""
    permission_classes = [IsAuthenticated, IsTeacherOrAdmin]
    queryset = ScoreAdjustment.objects.all()
    serializer_class = ScoreAdjustmentSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user

        if user.user_type == 'teacher':
            queryset = queryset.filter(
                Q(score__course_class__main_teacher=user) |
                Q(score__course_class__assistant_teachers=user)
            ).distinct()

        return queryset.select_related('score', 'score__student', 'adjusted_by', 'approved_by')

    def perform_create(self, serializer):
        serializer.save(adjusted_by=self.request.user)


class GradebookViewSet(viewsets.ModelViewSet):
    """记分册视图集"""
    permission_classes = [IsAuthenticated, IsTeacherOrAdmin]
    queryset = Gradebook.objects.all()

    def get_serializer_class(self):
        if self.action == 'create':
            return GradebookCreateSerializer
        return GradebookSerializer

    def get_queryset(self):
        """根据用户权限过滤查询集"""
        queryset = super().get_queryset()
        user = self.request.user
        course_class_id = self.request.query_params.get('course_class_id')

        logger.info(f"Gradebook查询 - 用户: {user.username}, 类型: {user.user_type}, course_class_id: {course_class_id}")

        # 处理course_class_id参数（可能是字符串）
        if course_class_id:
            try:
                course_class_id = int(course_class_id)
                queryset = queryset.filter(course_class_id=course_class_id)
                logger.info(f"过滤course_class_id={course_class_id}后，查询集数量: {queryset.count()}")
            except (ValueError, TypeError):
                logger.warning(f"无效的course_class_id: {course_class_id}")

        # 权限过滤
        if user.user_type == 'student':
            queryset = queryset.filter(student=user)
            logger.info(f"学生权限过滤后，查询集数量: {queryset.count()}")
        elif user.user_type == 'teacher':
            # 教师只能查看自己作为主讲教师或助教的班级的记分册
            queryset = queryset.filter(
                Q(course_class__main_teacher=user) |
                Q(course_class__assistant_teachers=user)
            ).distinct()
            logger.info(f"教师权限过滤后，查询集数量: {queryset.count()}")

        result = queryset.select_related('student', 'course_class', 'course_class__course').order_by('student__username')
        logger.info(f"最终查询集数量: {result.count()}")
        return result

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """导入记分册Excel"""
        from apps.scores.excel_handlers import GradebookExcelHandler

        serializer = ExcelUploadSerializer(data=request.data)
        if not serializer.is_valid():
            logger.error(f"导入记分册验证失败: {serializer.errors}")
            return Response({
                'success': False,
                'message': '数据验证失败',
                'errors': serializer.errors,
            }, status=status.HTTP_400_BAD_REQUEST)

        file = serializer.validated_data['file']
        course_class_id = serializer.validated_data['course_class_id']

        try:
            handler = GradebookExcelHandler()
            result = handler.parse_excel(file, course_class_id)

            logger.info(f"Excel解析结果: success={result.get('success')}, data_count={len(result.get('data', []))}, errors_count={len(result.get('errors', []))}")

            if not result['success']:
                return Response({
                    'success': False,
                    'message': 'Excel解析失败',
                    'errors': result['errors'],
                }, status=status.HTTP_400_BAD_REQUEST)

            # 检查是否有数据
            if not result.get('data') or len(result['data']) == 0:
                logger.warning("Excel解析成功但没有数据")
                return Response({
                    'success': False,
                    'message': 'Excel解析成功但没有找到有效数据',
                    'errors': result.get('errors', []) + ['未找到有效的数据行，请检查Excel文件格式'],
                }, status=status.HTTP_400_BAD_REQUEST)

            # 创建导入日志
            import_log = ScoreImportLog.objects.create(
                file_name=file.name,
                file_path='',
                course_class_id=course_class_id,
                imported_by=request.user,
                total_rows=len(result['data']),
                column_mapping=result.get('column_mapping', {}),
                status='processing'
            )

            # 导入数据
            logger.info(f"开始导入 {len(result['data'])} 条数据")
            import_result = handler.import_gradebooks(
                result['data'],
                course_class_id,
                request.user.id
            )
            logger.info(f"导入完成: 成功 {import_result['success']} 条, 失败 {import_result['failed']} 条")

            # 更新导入日志
            import_log.success_rows = import_result['success']
            import_log.failed_rows = import_result['failed']
            import_log.error_log = '\n'.join(import_result['errors'])
            import_log.status = 'completed' if import_result['failed'] == 0 else 'partial'
            import_log.completed_at = timezone.now()
            import_log.save()

            return Response({
                'success': True,
                'message': f'导入完成：成功 {import_result["success"]} 条，失败 {import_result["failed"]} 条',
                'data': {
                    'success_count': import_result['success'],
                    'failed_count': import_result['failed'],
                    'errors': import_result['errors'][:10],
                },
                'import_log_id': import_log.id,
            })

        except Exception as e:
            logger.error(f"导入记分册失败: {str(e)}", exc_info=True)
            return Response({
                'success': False,
                'message': f'导入失败: {str(e)}',
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """导出记分册Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from io import BytesIO

        course_class_id = request.query_params.get('course_class_id')
        if not course_class_id:
            return Response({'error': '请提供课程班级ID'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            course_class = CourseClass.objects.select_related('course').get(id=course_class_id)
            
            # 直接查询不过滤权限
            gradebooks = Gradebook.objects.filter(course_class_id=course_class_id).select_related(
                'student', 'course_class', 'course_class__course'
            ).order_by('student__username')

            logger.info(f"导出记分册 - course_class_id={course_class_id}, 记分册数量={gradebooks.count()}")

            wb = Workbook()
            ws = wb.active
            ws.title = '记分册'

            # 设置标题
            ws.merge_cells('A1:Q1')
            ws['A1'] = f'{course_class.course.course_name} - {course_class.class_name} 记分册'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            # 表头
            headers = [
                '学号', '姓名',
                '作业1', '作业2', '作业3', '作业4', '作业5',
                '实验1', '实验2',
                '考勤1', '考勤2', '考勤3', '考勤4', '考勤5',
                '复习笔记', '系统', '报告', '平时', '期末', '总评', '结论'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 数据行
            for row_idx, gradebook in enumerate(gradebooks, 3):
                try:
                    student_id = gradebook.student.student_profile.student_id
                except:
                    student_id = gradebook.student.employee_id or gradebook.student.username

                row_data = [
                    student_id,
                    gradebook.student.first_name or '',
                    gradebook.homework1 or '',
                    gradebook.homework2 or '',
                    gradebook.homework3 or '',
                    gradebook.homework4 or '',
                    gradebook.homework5 or '',
                    gradebook.experiment1 or '',
                    gradebook.experiment2 or '',
                    gradebook.attendance1 or '',
                    gradebook.attendance2 or '',
                    gradebook.attendance3 or '',
                    gradebook.attendance4 or '',
                    gradebook.attendance5 or '',
                    gradebook.review_note or '',
                    gradebook.system_score or '',
                    gradebook.report_score or '',
                    round(gradebook.usual_score) if gradebook.usual_score is not None else '',
                    round(gradebook.final_score) if gradebook.final_score is not None else '',
                    round(gradebook.total_score) if gradebook.total_score is not None else '',
                    gradebook.conclusion or '',
                ]
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col, value=value)

            # 设置列宽
            column_widths = [15, 10, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 10, 8, 8, 8, 8, 8]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width

            # 保存到BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"{course_class.course.course_name}_{course_class.class_name}_记分册.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except CourseClass.DoesNotExist:
            return Response({'error': '课程班级不存在'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"导出记分册失败: {str(e)}", exc_info=True)
            return Response({'error': f'导出失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='preview-graphics-gradebook')
    def preview_graphics_gradebook(self, request):
        """预览图形学记分册Excel（不导入）"""
        from apps.scores.excel_handlers import GradebookExcelHandler

        serializer = ExcelUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'error': '数据验证失败',
                'errors': serializer.errors,
            }, status=status.HTTP_400_BAD_REQUEST)

        file = serializer.validated_data['file']
        course_class_id = serializer.validated_data['course_class_id']

        try:
            handler = GradebookExcelHandler()
            result = handler.parse_excel(file, course_class_id)

            if not result['success']:
                return Response({
                    'success': False,
                    'error': 'Excel解析失败',
                    'errors': result['errors'],
                }, status=status.HTTP_400_BAD_REQUEST)

            # 计算预览数据
            preview_data = []

            def round_half_up(x):
                """四舍五入（标准数学舍入）"""
                if x is None:
                    return None
                import math
                return math.floor(x + 0.5)

            for item in result['data'][:20]:
                # 计算各平均值
                homework_avg = 0
                homework_count = 0
                homework_scores = []
                for key in ['homework1', 'homework2', 'homework3', 'homework4', 'homework5']:
                    if item.get(key) is not None:
                        homework_scores.append(item[key])
                        homework_count += 1
                if homework_scores:
                    homework_avg = sum(homework_scores) / len(homework_scores)

                experiment_avg = 0
                experiment_scores = []
                for key in ['experiment1', 'experiment2']:
                    if item.get(key) is not None:
                        experiment_scores.append(item[key])
                if experiment_scores:
                    experiment_avg = sum(experiment_scores) / len(experiment_scores)

                attendance_avg = 0
                attendance_scores = []
                for key in ['attendance1', 'attendance2', 'attendance3', 'attendance4', 'attendance5']:
                    if item.get(key) is not None:
                        attendance_scores.append(item[key])
                if attendance_scores:
                    attendance_avg = sum(attendance_scores) / len(attendance_scores)

                review_note = item.get('review_note') or 0
                system_score = item.get('system_score') or 0
                report_score = item.get('report_score') or 0

                # 计算平时成绩
                # 平时 = (AVERAGE(作业一到作业五)*0.1 + AVERAGE(实验一到实验二)*0.2 + AVERAGE(考勤1到考勤5)*0.05 + 复习笔记*0.05) / 0.4
                if homework_avg > 0 or experiment_avg > 0 or attendance_avg > 0 or review_note > 0:
                    usual_score = (homework_avg * 0.1 + experiment_avg * 0.2 + attendance_avg * 0.05 + review_note * 0.05) / 0.4
                    usual_score = round_half_up(usual_score)
                else:
                    usual_score = None

                # 计算期末成绩
                # 期末 = (系统 + 报告) / 2
                if system_score > 0 or report_score > 0:
                    final_score = (system_score + report_score) / 2
                else:
                    final_score = None

                # 计算总评
                # 总评 = 平时*0.4 + 期末*0.6
                if usual_score is not None and final_score is not None:
                    total_score = round_half_up(usual_score * 0.4 + final_score * 0.6)
                else:
                    total_score = None

                # 计算结论
                if total_score is not None:
                    if total_score >= 90:
                        conclusion = '优秀'
                    elif total_score >= 80:
                        conclusion = '良好'
                    elif total_score >= 70:
                        conclusion = '中等'
                    elif total_score >= 60:
                        conclusion = '及格'
                    else:
                        conclusion = '不及格'
                else:
                    conclusion = None

                preview_data.append({
                    'student_id': item.get('student_id'),
                    'student_name': item.get('student_name'),
                    'homework1': item.get('homework1'),
                    'homework2': item.get('homework2'),
                    'homework3': item.get('homework3'),
                    'homework4': item.get('homework4'),
                    'homework5': item.get('homework5'),
                    'experiment1': item.get('experiment1'),
                    'experiment2': item.get('experiment2'),
                    'attendance1': item.get('attendance1'),
                    'attendance2': item.get('attendance2'),
                    'attendance3': item.get('attendance3'),
                    'attendance4': item.get('attendance4'),
                    'attendance5': item.get('attendance5'),
                    'review_note': item.get('review_note'),
                    'system_score': system_score if system_score > 0 else None,
                    'report_score': report_score if report_score > 0 else None,
                    'usual_score': usual_score,
                    'final_score': round_half_up(final_score) if final_score is not None else None,
                    'total_score': total_score,
                    'conclusion': conclusion,
                })

            return Response({
                'success': True,
                'data': preview_data,
                'total_count': len(result['data']),
                'errors': result['errors'][:10],
            })

        except Exception as e:
            logger.error(f"预览图形学记分册失败: {str(e)}", exc_info=True)
            return Response({
                'success': False,
                'error': f'预览失败: {str(e)}',
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='import-graphics-gradebook')
    def import_graphics_gradebook(self, request):
        """导入图形学记分册Excel"""
        from apps.scores.excel_handlers import GradebookExcelHandler

        serializer = ExcelUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'message': '数据验证失败',
                'errors': serializer.errors,
            }, status=status.HTTP_400_BAD_REQUEST)

        file = serializer.validated_data['file']
        course_class_id = serializer.validated_data['course_class_id']

        try:
            handler = GradebookExcelHandler()
            result = handler.parse_excel(file, course_class_id)

            if not result['success']:
                return Response({
                    'success': False,
                    'message': 'Excel解析失败',
                    'errors': result['errors'],
                }, status=status.HTTP_400_NOT_FOUND)

            if not result.get('data') or len(result['data']) == 0:
                return Response({
                    'success': False,
                    'message': 'Excel解析成功但没有找到有效数据',
                    'errors': result.get('errors', []) + ['未找到有效的数据行，请检查Excel文件格式'],
                }, status=status.HTTP_400_BAD_REQUEST)

            # 创建导入日志
            import_log = ScoreImportLog.objects.create(
                file_name=file.name,
                file_path='',
                course_class_id=course_class_id,
                imported_by=request.user,
                total_rows=len(result['data']),
                column_mapping=result.get('column_mapping', {}),
                status='processing'
            )

            # 导入数据（使用标准的import_gradebooks方法）
            import_result = handler.import_gradebooks(
                result['data'],
                course_class_id,
                request.user.id
            )

            # 更新导入日志
            import_log.success_rows = import_result['success']
            import_log.failed_rows = import_result['failed']
            import_log.error_log = '\n'.join(import_result['errors'])
            import_log.status = 'completed' if import_result['failed'] == 0 else 'partial'
            import_log.completed_at = timezone.now()
            import_log.save()

            return Response({
                'success': True,
                'message': f'导入完成：成功 {import_result["success"]} 条，失败 {import_result["failed"]} 条',
                'data': {
                    'success_count': import_result['success'],
                    'failed_count': import_result['failed'],
                    'errors': import_result['errors'][:10],
                },
                'import_log_id': import_log.id,
            })

        except Exception as e:
            logger.error(f"导入图形学记分册失败: {str(e)}", exc_info=True)
            return Response({
                'success': False,
                'message': f'导入失败: {str(e)}',
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='graphics-classes')
    def graphics_classes(self, request):
        """获取图形学课程的班级列表"""
        from apps.courses.models import CourseClass, Course

        try:
            # 查找图形学课程
            courses = Course.objects.filter(course_name__icontains='图形学')
            course_ids = courses.values_list('id', flat=True)

            # 查找这些课程下的班级
            classes = CourseClass.objects.filter(course_id__in=course_ids).select_related('course', 'main_teacher')

            result = []
            for cls in classes:
                # 获取学生数量
                students_count = cls.students.count()
                # 获取记分册数量
                gradebook_count = cls.gradebooks.count()

                result.append({
                    'id': cls.id,
                    'course_id': cls.course_id,
                    'course_name': cls.course.course_name if cls.course else '',
                    'course_code': cls.course.course_code if cls.course else '',
                    'class_name': cls.class_name,
                    'main_teacher': cls.main_teacher_id,
                    'main_teacher_name': cls.main_teacher.first_name or cls.main_teacher.username if cls.main_teacher else '',
                    'students_count': students_count,
                    'gradebook_count': gradebook_count,
                })

            return Response({
                'success': True,
                'results': result
            })

        except Exception as e:
            logger.error(f"获取图形学班级列表失败: {str(e)}", exc_info=True)
            return Response({
                'success': False,
                'error': f'获取班级列表失败: {str(e)}',
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




